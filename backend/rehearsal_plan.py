"""Lokaler PDF-Import für Probenpläne und Zuordnung zu aktiven Mitarbeitern."""
from __future__ import annotations

import difflib
import os
import re
import unicodedata
from datetime import date, datetime, timedelta

import fitz
import openpyxl
from google import genai
from google.genai import types
from pydantic import BaseModel

GEMINI_MODEL = "gemini-flash-latest"

DEPARTMENT_SUFFIXES = {
    "dc", "tc", "ent", "gr", "deko", "spa", "wf", "waspo", "da", "fo",
    "ctm", "i&m", "tcb", "f&b", "tennisbar", "requi", "s&l", "spt",
}
GENERIC_PEOPLE = {
    "", '""', '"""', "tc's", "tcs", "alle die da sind", "alle",
}
DATE_RE = re.compile(
    r"(?:Mo|Di|Mi|Do|Fr|Sa|So)\.?\s*(\d{1,2})\.(\d{1,2})\.?",
    re.IGNORECASE,
)
TIME_RANGE_RE = re.compile(
    r"(\d{1,2})\s*[:.]\s*(\d{2})\s*-\s*(\d{1,2})\s*[:.]\s*(\d{2})"
)
TIME_ONLY_RE = re.compile(r"^\s*(\d{1,2})\s*[:.]\s*(\d{2})\s*$")
SUBSLOT_RE = re.compile(
    r"(.*?)\(\s*(\d{1,2})\s*[:.]\s*(\d{2})\s*-\s*"
    r"(\d{1,2})\s*[:.]\s*(\d{2})\s*\)",
    re.DOTALL,
)


class GeminiRehearsal(BaseModel):
    date: str
    start_time: str
    end_time: str
    location: str = ""
    activity: str
    show_code: str = ""
    participants_raw: str = ""
    choreographer_raw: str = ""
    end_inferred: bool = False


class GeminiRehearsalPlan(BaseModel):
    start_date: str
    end_date: str
    rehearsals: list[GeminiRehearsal]


GEMINI_PROMPT = """Du liest einen wöchentlichen Probenplan eines Hotel-Entertainment-Teams.
Extrahiere jede echte Probenzeile vollständig und in der Reihenfolge des PDFs.

Die Tabelle enthält typischerweise: Zeit | Wo | Was | Wer | Show | T&C.
- date: Datum der Tabellen-Sektion als YYYY-MM-DD.
- start_time/end_time: HH:MM. Geht das Ende über Mitternacht, bleibt z.B. 00:30.
- location: Inhalt aus „Wo“.
- activity: Inhalt aus „Was“.
- show_code: Inhalt aus „Show“.
- participants_raw: Namen aus „Wer“ exakt als kommaseparierter Text.
- choreographer_raw: Namen aus „T&C“ exakt als Text.
- end_inferred: nur true, wenn im PDF keine Endzeit steht. Dann 60 Minuten ansetzen.

Wichtig:
- Wiederholte Tabellenköpfe und reine Datumszeilen nicht als Probe ausgeben.
- Über Seitenumbrüche fortgesetzte Tage korrekt weiterführen.
- Persönliche Teilzeiten innerhalb einer Sammelprobe unbedingt im participants_raw
  direkt hinter dem jeweiligen Namen erhalten, z.B.
  „Livia (22:30-23:00), Dani, Sofia (23:00-23:30)“.
- Abteilungsanforderungen ohne konkrete Namen dürfen im participants_raw stehen bleiben.
- Tanzchoreografen nicht mit Teilnehmern vermischen.
- Nichts ergänzen, was nicht im PDF steht."""

SHOW_CODES = {
    "mr",
    "ny",
    "ror",
}
SHOW_NAMES = {
    "moulin rouge",
    "new york",
    "royals of rock",
}
PARTY_CODES = {
    "bn",
    "fys",
    "ww",
}
PARTY_NAMES = {
    "black night",
    "paradise on fire",
    "weiss weiss",
    "weiss-weiss",
}
REHEARSAL_ONLY_MARKERS = {
    "probe",
    "durchlauf",
    "meeting",
    "umzugshilfe",
    "spot",
}


def _norm(text: str) -> str:
    value = unicodedata.normalize("NFKD", text or "")
    value = "".join(char for char in value if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", value).strip().casefold()


def _clean_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _pdf_year(document: fitz.Document) -> int:
    creation = document.metadata.get("creationDate") or ""
    match = re.search(r"D:(\d{4})", creation)
    return int(match.group(1)) if match else datetime.now().year


def _parse_date(value: str, year: int) -> date | None:
    match = DATE_RE.search(value or "")
    if not match:
        return None
    return date(year, int(match.group(2)), int(match.group(1)))


def _clock(hour: str, minute: str) -> str:
    return f"{int(hour):02d}:{int(minute):02d}"


def _times(value: str) -> tuple[str, str, bool] | None:
    range_match = TIME_RANGE_RE.search(value or "")
    if range_match:
        return (
            _clock(range_match.group(1), range_match.group(2)),
            _clock(range_match.group(3), range_match.group(4)),
            False,
        )
    only = TIME_ONLY_RE.match(value or "")
    if only:
        start = datetime.strptime(
            _clock(only.group(1), only.group(2)), "%H:%M"
        )
        # Einzelzeiten werden aus Sicherheitsgründen als einstündige Probe behandelt.
        end = start + timedelta(hours=1)
        return start.strftime("%H:%M"), end.strftime("%H:%M"), True
    return None


def _strip_participant_count(text: str) -> str:
    return re.sub(
        r"\(\s*\d+\s*(?:teiln(?:e|eh)mer|teilnhemer|teilnehmer)\s*\)",
        "",
        text,
        flags=re.IGNORECASE,
    )


def _candidate_name(raw: str) -> str | None:
    value = _clean_text(raw).strip(" ,;")
    value = re.sub(r"^\d+x\s+.*$", "", value, flags=re.IGNORECASE)
    value = re.sub(r"\([^)]*\)", "", value).strip()
    if _norm(value) in GENERIC_PEOPLE:
        return None
    parts = value.split()
    while parts and _norm(parts[-1]) in DEPARTMENT_SUFFIXES:
        parts.pop()
    value = " ".join(parts).strip(" ,;")
    if not value or _norm(value) in GENERIC_PEOPLE:
        return None
    return value


def _match_active(raw_name: str, active_people: list[str]) -> str | None:
    candidate = _candidate_name(raw_name)
    if not candidate:
        return None
    normalized = _norm(candidate)
    exact = { _norm(name): name for name in active_people }
    if normalized in exact:
        return exact[normalized]
    # Tippfehler wie Killian/Kilian aus dem PDF vorsichtig auflösen.
    # Nur sehr sichere Tippvarianten automatisch zusammenführen (z. B.
    # "Killian" -> "Kilian"). Ähnliche, aber eigenständige Namen wie
    # "Melina" und "Selina" müssen in der Prüfansicht bestätigt werden.
    close = difflib.get_close_matches(normalized, list(exact), n=1, cutoff=0.90)
    return exact[close[0]] if close else None


def _split_names(text: str) -> list[str]:
    cleaned = _strip_participant_count(_clean_text(text))
    parts = re.split(r"\s*,\s*|\s+&\s+", cleaned)
    return [
        candidate
        for part in parts
        if (candidate := _candidate_name(part))
    ]


def _people_for_row(
    participants_raw: str,
    choreographer_raw: str,
    start_time: str,
    end_time: str,
    active_people: list[str],
) -> list[dict]:
    people: list[dict] = []
    normalized_participants = _strip_participant_count(
        (participants_raw or "").replace("\n", " ")
    )
    subslots = list(SUBSLOT_RE.finditer(normalized_participants))
    consumed_until = 0
    if subslots:
        for match in subslots:
            names_text = match.group(1)
            sub_start = _clock(match.group(2), match.group(3))
            sub_end = _clock(match.group(4), match.group(5))
            for raw_name in _split_names(names_text):
                people.append({
                    "raw_name": raw_name,
                    "person_name": _match_active(raw_name, active_people),
                    "role": "participant",
                    "start_time": sub_start,
                    "end_time": sub_end,
                })
            consumed_until = match.end()
        trailing = normalized_participants[consumed_until:].strip(" ,;")
        for raw_name in _split_names(trailing):
            people.append({
                "raw_name": raw_name,
                "person_name": _match_active(raw_name, active_people),
                "role": "participant",
                "start_time": start_time,
                "end_time": end_time,
            })
    else:
        for raw_name in _split_names(normalized_participants):
            people.append({
                "raw_name": raw_name,
                "person_name": _match_active(raw_name, active_people),
                "role": "participant",
                "start_time": start_time,
                "end_time": end_time,
            })

    for raw_name in _split_names(choreographer_raw):
        people.append({
            "raw_name": raw_name,
            "person_name": _match_active(raw_name, active_people),
            "role": "choreographer",
            "start_time": start_time,
            "end_time": end_time,
        })
    return people


def normalize_rehearsals(rehearsals: list[dict], active_people: list[str]) -> list[dict]:
    result = []
    for row in rehearsals:
        start_time = row["start_time"]
        end_time = row["end_time"]
        item = {
            "date": row["date"],
            "start_time": start_time,
            "end_time": end_time,
            "location": _clean_text(row.get("location")),
            "activity": _clean_text(row.get("activity")) or "Probe",
            "show_code": _clean_text(row.get("show_code")),
            "participants_raw": _clean_text(row.get("participants_raw")),
            "choreographer_raw": _clean_text(row.get("choreographer_raw")),
            "end_inferred": bool(row.get("end_inferred")),
        }
        item["people"] = _people_for_row(
            item["participants_raw"],
            item["choreographer_raw"],
            start_time,
            end_time,
            active_people,
        )
        result.append(item)
    return result


def is_show_event(row: dict) -> bool:
    """Prüft, ob ein einzelner Probenplan-Termin ein echter Showabend ist.

    Ein abendlicher Eintrag mit Show-Kennung gilt als Show. Bekannte Partyformate
    sowie ausdrücklich als Probe/Durchlauf markierte Termine zählen nicht.
    """
    activity = _norm(row.get("activity") or "")
    show_code = _norm(row.get("show_code") or "").strip('"')
    start_time = row.get("start_time") or ""
    is_evening = start_time >= "20:00"
    is_party = (
        show_code in PARTY_CODES
        or any(name in activity for name in PARTY_NAMES)
    )
    is_rehearsal_only = any(
        marker in activity for marker in REHEARSAL_ONLY_MARKERS
    )
    is_explicit_show = is_evening and "show" in activity
    is_known_evening_show = (
        is_evening
        and (
            show_code in SHOW_CODES
            or any(name in activity for name in SHOW_NAMES)
        )
    )
    is_coded_evening_show = (
        is_evening
        and bool(show_code)
        and not is_party
        and not is_rehearsal_only
    )
    return not is_party and not is_rehearsal_only and (
        is_explicit_show or is_known_evening_show or is_coded_evening_show
    )


def detect_show_dates(rehearsals: list[dict]) -> set[str]:
    """Erkennt echte Showabende aus dem Probenplan (siehe `is_show_event`)."""
    result: set[str] = set()
    for raw in rehearsals:
        row = dict(raw)
        if is_show_event(row):
            result.add(row["date"])
    return result


def extract_pdf_with_gemini(
    pdf_bytes: bytes,
    active_people: list[str],
    source_filename: str | None = None,
    api_key: str | None = None,
) -> dict:
    """Liest einen Probenplan mit Gemini und normalisiert Namen lokal."""
    key = api_key or os.environ.get("GEMINI_API_KEY")
    if not key:
        raise RuntimeError("Kein Gemini API Key vorhanden.")

    document = fitz.open(stream=pdf_bytes, filetype="pdf")
    document_year = _pdf_year(document)
    document.close()
    client = genai.Client(api_key=key)
    response = client.models.generate_content(
        model=GEMINI_MODEL,
        contents=[
            types.Part.from_bytes(data=pdf_bytes, mime_type="application/pdf"),
            (
                "Extrahiere diesen Probenplan vollständig. "
                f"Das verbindliche Dokumentjahr ist {document_year}."
            ),
        ],
        config=types.GenerateContentConfig(
            system_instruction=GEMINI_PROMPT,
            response_mime_type="application/json",
            response_schema=GeminiRehearsalPlan,
        ),
    )
    parsed: GeminiRehearsalPlan | None = response.parsed
    if parsed is None or not parsed.rehearsals:
        raise RuntimeError(
            "Gemini hat keine gültigen Probenzeilen zurückgegeben."
        )

    rows: list[dict] = []
    warnings: list[str] = []
    for rehearsal in parsed.rehearsals:
        item = rehearsal.model_dump()
        parsed_date = datetime.strptime(item["date"], "%Y-%m-%d").date()
        item["date"] = parsed_date.replace(year=document_year).isoformat()
        parsed_range = _times(
            f"{item['start_time']}-{item['end_time']}"
        )
        if parsed_range is None:
            raise ValueError(
                f"Ungültige Probenzeit: {item['start_time']}–{item['end_time']}"
            )
        item["start_time"], item["end_time"], _ = parsed_range
        rows.append(item)
        if item["end_inferred"]:
            warnings.append(
                f"{datetime.strptime(item['date'], '%Y-%m-%d').strftime('%d.%m.')}: "
                f"Ende für „{item['activity']}“ wurde mit 60 Minuten angesetzt."
            )

    normalized = normalize_rehearsals(rows, active_people)
    dates = [
        datetime.strptime(row["date"], "%Y-%m-%d").date()
        for row in normalized
    ]
    return {
        "start_date": min(dates).isoformat(),
        "end_date": max(dates).isoformat(),
        "source_filename": source_filename,
        "rehearsals": normalized,
        "warnings": warnings,
        "extraction_method": "gemini",
    }


# --- Excel-Import -----------------------------------------------------------

# Wochenblätter heißen "03.08-09.08", "29.06 - 05.07" oder "09.02-15.02.2026".
SHEET_RANGE_RE = re.compile(
    r"(\d{1,2})[.,](\d{1,2})\s*-\s*(\d{1,2})[.,](\d{1,2})(?:[.,](\d{4}))?"
)
WEEKDAY_INDEX = {"mo": 0, "di": 1, "mi": 2, "do": 3, "fr": 4, "sa": 5, "so": 6}
WEEKDAY_RE = re.compile(r"^\s*(mo|di|mi|do|fr|sa|so)\b", re.IGNORECASE)
# Kopfzeile eines Tagesblocks: Zeit | Wo | Was | Wer | Show | T&C
HEADER_TOKENS = {"zeit", "wo", "was", "wer", "show"}


def list_sheets(source) -> list[str]:
    """Nur echte Wochenblätter - "Vorlage", "Diagramm1", "FDK" usw. fliegen raus."""
    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    try:
        return [name for name in wb.sheetnames if SHEET_RANGE_RE.search(name)]
    finally:
        wb.close()


def sheet_week_start(sheet_name: str, reference: date | None = None) -> date | None:
    """Wochenstart aus dem Blattnamen. Ohne Jahresangabe wird das Jahr gewählt,
    dessen Datum dem Referenztag (heute) am nächsten liegt."""
    match = SHEET_RANGE_RE.search(sheet_name or "")
    if not match:
        return None
    day, month = int(match.group(1)), int(match.group(2))
    if match.group(5):
        try:
            return date(int(match.group(5)), month, day)
        except ValueError:
            return None
    today = reference or date.today()
    candidates = []
    for year in (today.year - 1, today.year, today.year + 1):
        try:
            candidates.append(date(year, month, day))
        except ValueError:
            continue
    return min(candidates, key=lambda d: abs((d - today).days)) if candidates else None


def extract_xlsx(
    source,
    active_people: list[str],
    sheet_name: str | None = None,
    source_filename: str | None = None,
) -> dict:
    """Liest ein Wochenblatt des Probenplans aus Excel.

    Die Datumszeilen im Blatt ("Di. 04.09") sind in echten Dateien oft aus einer
    anderen Woche kopiert und tragen den falschen Monat. Verlässlich ist nur der
    Wochentag - das Datum kommt deshalb aus dem Blattnamen plus Wochentag-Versatz.
    Abweichungen werden als Warnung gemeldet, statt sie stillschweigend zu übernehmen.
    """
    wb = openpyxl.load_workbook(source, data_only=True)
    try:
        if sheet_name is None:
            sheets = [n for n in wb.sheetnames if SHEET_RANGE_RE.search(n)]
            if not sheets:
                raise ValueError("Die Datei enthält kein erkennbares Wochenblatt.")
            sheet_name = sheets[-1]
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Blatt '{sheet_name}' fehlt in der Datei.")
        ws = wb[sheet_name]

        week_start = sheet_week_start(sheet_name)
        if week_start is None:
            raise ValueError(
                f"Aus dem Blattnamen '{sheet_name}' liess sich kein Wochenstart lesen."
            )

        rows: list[dict] = []
        warnings: list[str] = []
        mismatches: set[str] = set()
        current: date | None = None

        for row in ws.iter_rows(min_row=1, max_col=6, values_only=True):
            cells = [_clean_text(value) for value in row] + [""] * 6
            first = cells[0]

            weekday = WEEKDAY_RE.match(first)
            if weekday and not _times(first):
                index = WEEKDAY_INDEX[weekday.group(1).casefold()]
                current = week_start + timedelta(days=index)
                printed = _parse_date(first, week_start.year)
                if printed and printed != current:
                    mismatches.add(f"{first} -> {current.strftime('%d.%m.')}")
                continue

            if first.casefold() in HEADER_TOKENS or cells[1].casefold() == "wo":
                continue
            if current is None:
                continue

            times = _times(first)
            if times is None:
                continue
            start_time, end_time, inferred = times
            activity = cells[2] or cells[1]
            if not activity:
                continue

            rows.append({
                "date": current.isoformat(),
                "start_time": start_time,
                "end_time": end_time,
                "location": cells[1],
                "activity": activity,
                "show_code": cells[4],
                "participants_raw": cells[3],
                "choreographer_raw": cells[5],
                "end_inferred": inferred,
            })
            if inferred:
                warnings.append(
                    f"{current.strftime('%d.%m.')}: Ende für „{activity}“ "
                    "wurde mit 60 Minuten angesetzt."
                )
    finally:
        wb.close()

    if not rows:
        raise ValueError(f"Im Blatt '{sheet_name}' wurden keine Probenzeilen erkannt.")

    if mismatches:
        warnings.insert(0, (
            "Im Blatt stehen abweichende Datumsangaben (vermutlich aus einer anderen "
            "Woche kopiert). Verwendet wurde der Wochentag aus dem Blattnamen: "
            + ", ".join(sorted(mismatches))
        ))

    normalized = normalize_rehearsals(rows, active_people)
    dates = [datetime.strptime(row["date"], "%Y-%m-%d").date() for row in normalized]
    return {
        "start_date": min(dates).isoformat(),
        "end_date": max(dates).isoformat(),
        "source_filename": source_filename,
        "sheet_name": sheet_name,
        "rehearsals": normalized,
        "warnings": warnings,
        "extraction_method": "excel",
    }


def extract_pdf(
    pdf_bytes: bytes,
    active_people: list[str],
    source_filename: str | None = None,
) -> dict:
    document = fitz.open(stream=pdf_bytes, filetype="pdf")
    year = _pdf_year(document)
    current_date: date | None = None
    rehearsals: list[dict] = []
    warnings: list[str] = []

    for page in document:
        for table in page.find_tables().tables:
            for raw_row in table.extract():
                row = list(raw_row) + [None] * (6 - len(raw_row))
                cells = [_clean_text(value) for value in row[:6]]
                row_date = _parse_date(cells[0], year)
                if row_date:
                    current_date = row_date
                    continue
                if _norm(cells[0]) == "zeit" or not cells[0]:
                    continue
                parsed_times = _times(cells[0])
                if parsed_times is None or current_date is None:
                    continue
                start_time, end_time, inferred = parsed_times
                activity = cells[2] or cells[1] or "Probe"
                # Einzelzeile „21:45 | Clubtanz Probe mit Sofia“ liegt wegen der
                # Tabellenformatierung in der Orts-Spalte.
                location = cells[1]
                participants_raw = cells[3]
                choreographer_raw = cells[5]
                if not cells[2] and "probe" in _norm(cells[1]):
                    activity = cells[1]
                    location = ""
                    name_match = re.search(r"\bmit\s+(.+?)(?:<3)?$", activity, re.IGNORECASE)
                    if name_match and not choreographer_raw:
                        choreographer_raw = name_match.group(1).strip()
                item = {
                    "date": current_date.isoformat(),
                    "start_time": start_time,
                    "end_time": end_time,
                    "location": location,
                    "activity": activity,
                    "show_code": cells[4],
                    "participants_raw": participants_raw,
                    "choreographer_raw": choreographer_raw,
                    "end_inferred": inferred,
                }
                rehearsals.append(item)
                if inferred:
                    warnings.append(
                        f"{current_date.strftime('%d.%m.')}: Ende für „{activity}“ "
                        "war nicht angegeben und wurde mit 60 Minuten angesetzt."
                    )
    document.close()
    normalized = normalize_rehearsals(rehearsals, active_people)
    dates = [datetime.strptime(row["date"], "%Y-%m-%d").date() for row in normalized]
    if not dates:
        raise ValueError("Im PDF wurden keine Probenzeilen erkannt.")
    start = min(dates)
    end = max(dates)
    return {
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "source_filename": source_filename,
        "rehearsals": normalized,
        "warnings": warnings,
        "extraction_method": "local",
    }
