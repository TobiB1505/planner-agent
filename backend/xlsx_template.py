"""Erzeugt neue Wochenpläne aus der echten Excel-Vorlage (Formatierung bleibt exakt erhalten)
und liest ausgefüllte Excel-Dienstpläne deterministisch wieder ein (kein LLM nötig)."""
from __future__ import annotations

import re
from datetime import date, datetime, timedelta

import openpyxl

from . import grid
from . import planning_rules
from . import template_spec


def _normalize_label(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[.:/+\-]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


# Aus template_spec.py abgeleitet (Tobis Zeile-für-Zeile-Erklärung der echten Vorlage).
# Abteilungs-Zeilen werden beim Import separat behandelt, damit Abteilungsnamen nicht
# versehentlich als neue Mitarbeiter angelegt werden.
DEPARTMENT_LABELS = {_normalize_label(name) for name in template_spec.department_labels()}
# AP9: zentrale Quelle in template_spec.py - vorher eine eigene, wertgleiche Kopie hier.
DEPARTMENT_TOKENS = template_spec.DEPARTMENT_TOKENS
GENERIC_DETAIL_LABELS = {"wann wo", "wer wann wo", "wer"}
DETAIL_ROW_CATEGORIES = {
    _normalize_label(grid._normalize_category_name(name))
    for name in (
        "Aufbau",
        "Abbau",
        "Reminders",
        "Nachmittag/Abend Extras",
        "18:00 Saunaaufguss",
        "00:30-03:00 NITE CLUB",
        "MOD",
    )
}


def _canonical_category(label: str) -> str | None:
    """Ordnet Varianten aus den beiden echten Excel-Dateien der zentralen Spezifikation zu."""
    norm = _normalize_label(label)
    if not norm:
        return None
    best: tuple[int, str] | None = None
    for name, spec in template_spec.ROWS.items():
        spec_norm = _normalize_label(name)
        if norm == spec_norm:
            score = 1000
        elif spec_norm in norm or norm in spec_norm:
            score = min(len(norm), len(spec_norm))
        else:
            continue
        category = (
            name
            if spec["kind"] == template_spec.ABSENCE
            else grid._normalize_category_name(spec.get("family") or name)
        )
        if best is None or score > best[0]:
            best = (score, category)
    return best[1] if best else None


def _is_department_value(text: str) -> bool:
    return text.strip().upper() in DEPARTMENT_TOKENS


def _clean_person_token(token: str) -> str | None:
    """Bereinigt typische Zusätze der echten Excel-Pläne, ohne daraus neue MA zu machen."""
    token = re.sub(r"\s+", " ", token or "").strip(" \t\r\n()")
    token = re.sub(r"\s*\((?:GT|Dienst[^)]*)\)\s*$", "", token, flags=re.IGNORECASE)
    token = re.sub(r"\s+für\s+.+$", "", token, flags=re.IGNORECASE)
    token = token.strip(" \t\r\n():")
    if not token or token.casefold() in {"-", "keine", "kein", "niemand"}:
        return None
    corrections = {
        "manü": "Manu",
    }
    return corrections.get(token.casefold(), token)


def _split_people(text: str) -> list[str]:
    """Trennt mehrere MA und entfernt Klammer-/Hinweis-Schreibweisen aus den Vorlagen."""
    value = re.sub(
        r"^\s*\(Gäste\s+vs\.?\s+Robins\s+BVB\)\s*",
        "",
        text or "",
        flags=re.IGNORECASE,
    )
    value = value.strip()
    if value.startswith("(") and value.endswith(")"):
        value = value[1:-1]
    value = re.sub(r"\((?:GT|Dienst[^)]*)\)", "", value, flags=re.IGNORECASE)
    # Ein Name in nachgestellten Klammern ist in den Altplänen nur Vertretungs-/Hinweistext
    # (z.B. "Nick (Kilian)") und keine zweite feste Einteilung.
    value = re.sub(r"(?<=\w)\s*\([^,]+\)\s*$", "", value)

    people: list[str] = []
    for raw in re.split(r"\s*[,+/]\s*", value):
        person = _clean_person_token(raw)
        if person and not _is_department_value(person):
            people.append(person)
    return people


def _person_cell_parts(category: str, label: str, text: str) -> tuple[str | None, list[str]]:
    """Liest je Zeilentyp nur den echten MA-Anteil einer Zelle.

    Rückgabe: (Subkategorie/Uhrzeit, erkannte MA). Abteilungswerte ergeben bewusst keine
    Person und bleiben über ``raw_text`` trotzdem vollständig im Archiv erhalten.
    """
    normalized_category = grid._normalize_category_name(category)
    value = text.strip()
    subcategory: str | None = None
    person_part = value

    if normalized_category == "OPS + WP":
        person_part = re.sub(r"^\s*\d{1,2}:\d{2}\s*", "", value)
        person_part = re.sub(r"\b(?:OPS|WP|AL)\b", " ", person_part, flags=re.IGNORECASE)
        person_part = re.sub(r"\s+", " ", person_part).strip(" +")

    elif normalized_category == "Aperitif":
        lines = [part.strip() for part in re.split(r"[\r\n]+", value) if part.strip()]
        # Ort, Uhrzeit und Künstler sind Inhalt; nur eine vierte/letzte Zeile ist der MA
        # bzw. die zuständige Abteilung. Bei nur drei Zeilen fehlt die MA-Zuweisung.
        if len(lines) < 4:
            return None, []
        subcategory = " | ".join(lines[:-1])
        person_part = lines[-1]

    elif normalized_category == "An + Abreise-Dienst":
        if ":" in value:
            subcategory, person_part = (part.strip() for part in value.rsplit(":", 1))

    elif normalized_category == "Sportprogramm":
        marker = ""
        if re.match(r"^\s*\(Gäste\s+vs\.?\s+Robins\s+BVB\)", value, re.IGNORECASE):
            marker = " (Gäste vs Robins BVB)"
        if "Softsport" in label:
            parts = re.split(r"\s+(?:\||I)\s+", value, maxsplit=1)
            if len(parts) == 2:
                activity, person_part = parts
                subcategory = f"{label} {activity.strip()}".strip()
            else:
                subcategory = label
        else:
            subcategory = f"{label}{marker}"
            parts = re.split(r"\s+(?:\||I)\s+", person_part, maxsplit=1)
            if len(parts) == 2:
                person_part = parts[1]

    elif normalized_category == "Kochdienste":
        subcategory = label

    return subcategory, _split_people(person_part)


def _sheet_dates(ws) -> tuple[list[tuple[int, date]], int]:
    """Findet die Datumszeile sowohl in alten Ein-Zeilen- als auch neuen Tag/Datum-Köpfen."""
    best_dates: list[tuple[int, date]] = []
    best_row = 1
    for row in range(1, min(ws.max_row, 3) + 1):
        found: list[tuple[int, date]] = []
        for col in range(2, 9):
            value = ws.cell(row=row, column=col).value
            parsed: date | None = None
            if isinstance(value, datetime):
                parsed = value.date()
            elif isinstance(value, date):
                parsed = value
            elif value:
                match = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", str(value))
                if match:
                    parsed = date(int(match.group(3)), int(match.group(2)), int(match.group(1)))
            if parsed:
                found.append((col, parsed))
        if len(found) > len(best_dates):
            best_dates = found
            best_row = row
    return best_dates, best_row


def read_template_content(
    template_path: str,
    sheet_name: str,
    new_start: date,
) -> list[dict]:
    """Liest feste Programm-/Infofelder aus Woche A oder B und legt sie auf neue Datumswerte.

    Personen-, Frei-/Urlaubs- und MOD-Zeilen werden bewusst nicht übernommen. Diese werden
    im neuen Plan automatisch verteilt bzw. manuell gepflegt.
    """
    wb = openpyxl.load_workbook(template_path, data_only=True)
    try:
        ws = wb[sheet_name]
        _, date_row = _sheet_dates(ws)
        rows: list[dict] = []
        parent_category: str | None = None

        for row in range(date_row + 1, ws.max_row + 1):
            raw_label = ws.cell(row=row, column=1).value
            if not raw_label:
                continue
            label = str(raw_label).strip()
            norm_label = _normalize_label(label)
            direct_category = _canonical_category(label)
            if direct_category:
                parent_category = direct_category
                category = direct_category
            elif norm_label in GENERIC_DETAIL_LABELS and parent_category:
                category = parent_category
            else:
                continue

            kind = grid.category_kind(category)
            if kind not in (template_spec.CONTENT, template_spec.DEPARTMENT):
                continue
            if category == grid._normalize_category_name("Specials"):
                continue

            for day_index, col in enumerate(range(2, 9)):
                value = ws.cell(row=row, column=col).value
                if value is None:
                    continue
                text = str(value).strip()
                if not text:
                    continue
                rows.append({
                    "date": (new_start + timedelta(days=day_index)).isoformat(),
                    "category": category,
                    "subcategory": None,
                    "person": None,
                    "raw_text": text,
                })

        return rows
    finally:
        wb.close()


def list_week_sheets(path: str) -> list[str]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        return [s for s in wb.sheetnames if s.strip().lower() != "mitarbeiter"]
    finally:
        wb.close()


def _row_labels(ws) -> dict[int, str]:
    labels = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v:
            labels[r] = str(v).strip()
    return labels


def generate_week_xlsx(
    template_path: str, sheet_name: str, new_start: date,
    draft_rows: list[dict], absences: list[dict], output_path: str,
) -> str:
    """Dupliziert das gewählte Vorlagen-Blatt (Farben/Merges/Spaltenbreiten bleiben exakt
    erhalten, da openpyxl dieselbe Formatierung kopiert statt sie nachzubilden), leert alle
    Datenzellen, aktualisiert die Datumszeile und trägt Fairness-Vorschläge + Abwesenheiten
    an den passenden Zeilen ein. Inhalte-Zeilen (Show/Party, MOD, ...) bleiben leer."""
    wb = openpyxl.load_workbook(template_path)
    try:
        src = wb[sheet_name]

        new_title = f"KW{new_start.isocalendar()[1]}_{new_start.strftime('%d.%m')}"[:31]
        title, n = new_title, 1
        while title in wb.sheetnames:
            n += 1
            title = f"{new_title}_{n}"[:31]

        ws = wb.copy_worksheet(src)
        ws.title = title

        labels = _row_labels(ws)
        day_col = {(new_start + timedelta(days=i)).isoformat(): 2 + i for i in range(7)}

        _, source_date_row = _sheet_dates(ws)
        two_row_header = (
            _normalize_label(str(ws.cell(row=1, column=1).value or "")) == "tag"
            and _normalize_label(str(ws.cell(row=2, column=1).value or "")) == "datum"
        )
        date_row = 2 if two_row_header else source_date_row
        for i in range(7):
            col = 2 + i
            d = new_start + timedelta(days=i)
            if two_row_header:
                ws.cell(row=1, column=col).value = grid.WEEKDAY_NAMES_DE[i]
                ws.cell(row=2, column=col).value = d
            else:
                ws.cell(row=date_row, column=col).value = (
                    f"{grid.WEEKDAY_NAMES_DE[i][:2]}, {d.strftime('%d.%m.%Y')}"
                )

        for r in labels:
            if r <= date_row:
                continue
            for c in range(2, 9):
                ws.cell(row=r, column=c).value = None

        slot_rows: dict[tuple[str, str], int] = {}
        norm_rows: dict[str, int] = {}
        for r, label in labels.items():
            if r <= date_row:
                continue
            row_category = _canonical_category(label)
            if row_category in grid.FAMILY_SLOT_LABELS:
                slot = grid.slot_key_for_category(row_category, label)
                slot_rows.setdefault((row_category, slot), r)
            norm_rows.setdefault(_normalize_label(label), r)

        def find_row(category: str, subcategory: str | None) -> int | None:
            # Zeitbasierte Zuordnung nur für Kategorien, die wir selbst in mehrere Zeitschienen-
            # Zeilen aufteilen (Sportprogramm, Kochdienste) - sonst könnten generische Uhrzeiten
            # wie "11:00" fälschlich in eine unabhängige Zeile derselben Uhrzeit rutschen
            # (z.B. OPS/WP -> Sportprogramm-Zeile "11:00 BVB").
            if category in ("Sportprogramm", "Kochdienste") and subcategory:
                sk = grid.slot_key_for_category(category, subcategory)
                if (category, sk) in slot_rows:
                    return slot_rows[(category, sk)]
            norm_cat = _normalize_label(category)
            if not norm_cat:
                return None
            header_row = norm_rows.get(norm_cat)
            if header_row is None:
                for norm_label, r in norm_rows.items():
                    if norm_cat in norm_label or norm_label in norm_cat:
                        header_row = r
                        break
            if header_row is None:
                return None

            if norm_cat in DETAIL_ROW_CATEGORIES:
                for detail_row in range(header_row + 1, min(header_row + 3, ws.max_row + 1)):
                    detail_label = _normalize_label(labels.get(detail_row, ""))
                    if detail_label in GENERIC_DETAIL_LABELS:
                        return detail_row
            return header_row

        absence_row = {label: r for r, label in labels.items() if label in grid.ABSENCE_ROWS}
        cell_names: dict[tuple[int, int], list[str]] = {}
        cell_text: dict[tuple[int, int], str] = {}
        special_bvb_cells: set[tuple[int, int]] = set()

        for a in absences:
            row_label = "Frei" if a["type"] == "Frei" else "Urlaub/Krank"
            r = absence_row.get(row_label)
            col = day_col.get(a["date"])
            if r and col:
                cell_names.setdefault((r, col), []).append(a["person"])

        for a in draft_rows:
            r = find_row(a["category"], a.get("subcategory"))
            col = day_col.get(a["date"])
            if r and col and a.get("person"):
                if planning_rules.is_guests_vs_robins_bvb(
                    a["category"], a.get("subcategory"), a.get("date")
                ):
                    special_bvb_cells.add((r, col))
                cell_names.setdefault((r, col), []).append(
                    grid.format_person_assignment(
                        a["category"], a.get("subcategory"), a["person"]
                    )
                )
            elif r and col and a.get("raw_text"):
                cell_text[(r, col)] = str(a["raw_text"]).strip()

        for (r, c), text in cell_text.items():
            ws.cell(row=r, column=c).value = text

        for (r, c), names in cell_names.items():
            value = ", ".join(names)
            if (r, c) in special_bvb_cells:
                value = f"(Gäste vs Robins BVB)\n{value}"
            ws.cell(row=r, column=c).value = value

        wb.save(output_path)
        return output_path
    finally:
        wb.close()


def extract_from_xlsx(path: str, sheet_name: str | None = None) -> dict:
    """Liest einen ausgefüllten Dienstplan direkt aus den Zellwerten (kein LLM, deterministisch).
    Liefert dasselbe Format wie extraction.extract_dienstplan (kw, start_date, end_date,
    assignments[], absences[]), damit der bestehende Prüf-/Speicherfluss wiederverwendet wird."""
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        if sheet_name is None:
            candidates = [s for s in wb.sheetnames if s.strip().lower() != "mitarbeiter"]
            sheet_name = candidates[-1]
        ws = wb[sheet_name]

        dates, date_row = _sheet_dates(ws)
        if not dates:
            raise ValueError("Konnte die Datumszeile nicht lesen.")

        start_date = min(d for _, d in dates)
        end_date = max(d for _, d in dates)

        assignments: list[dict] = []
        absences: list[dict] = []
        parent_category: str | None = None

        for r in range(date_row + 1, ws.max_row + 1):
            raw_label = ws.cell(row=r, column=1).value
            if not raw_label:
                continue
            label = str(raw_label).strip()
            if not label:
                continue
            norm_label = _normalize_label(label)
            direct_category = _canonical_category(label)
            if direct_category:
                parent_category = direct_category
            category = (
                parent_category
                if norm_label in GENERIC_DETAIL_LABELS and parent_category
                else direct_category or label
            )
            kind = grid.category_kind(category)

            for col, d in dates:
                val = ws.cell(row=r, column=col).value
                if val is None:
                    continue
                text = str(val).strip()
                if not text or text == "-":
                    continue

                if category in grid.ABSENCE_ROWS:
                    typ = category
                    for name in _split_people(text):
                        absences.append({"date": d.isoformat(), "person": name, "type": typ})
                    continue

                if kind in (template_spec.CONTENT, template_spec.MANUAL):
                    # Shows, Meetings, Mottos, Erinnerungen und MOD sind wichtige Planinhalte,
                    # aber keine Mitarbeiter. Sie werden als Freitext archiviert.
                    assignments.append({
                        "date": d.isoformat(), "category": category, "subcategory": None,
                        "person": None, "raw_text": text,
                    })
                    continue

                if kind == template_spec.DEPARTMENT or norm_label in DEPARTMENT_LABELS:
                    # Abteilungs-Zeile (z.B. Abbauhilfe, 18 Uhr LEDs): vollständig archivieren,
                    # aber niemals als Mitarbeiter in den aktiven Pool übernehmen.
                    assignments.append({
                        "date": d.isoformat(), "category": category, "subcategory": None,
                        "person": None, "raw_text": text,
                    })
                    continue

                subcat, people = _person_cell_parts(category, label, text)
                if not people:
                    # "keine", reine Abteilungswerte oder ein Aperitif ohne MA bleiben als
                    # Information erhalten, erzeugen aber keine Person.
                    assignments.append({
                        "date": d.isoformat(), "category": category, "subcategory": subcat,
                        "person": None, "raw_text": text,
                    })
                    continue
                for name in people:
                    assignments.append({
                        "date": d.isoformat(), "category": category, "subcategory": subcat,
                        "person": name, "raw_text": text,
                    })

        return {
            "kw": start_date.isocalendar()[1],
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "assignments": assignments,
            "absences": absences,
        }
    finally:
        wb.close()
