"""Tests für die Excel-Programmvorlagen unter backend/resources/templates/.

Öffnet Vorlagen ausschliesslich lesend (read_only=True) und vergleicht vor/
nach Byte für Byte, damit sichergestellt ist, dass Tests sie nie verändern.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from backend.config import paths as paths_module

REQUIRED_TEMPLATES = [
    paths_module.WEEK_A_TEMPLATE_PATH,
    paths_module.WEEK_B_TEMPLATE_PATH,
    paths_module.ARTIST_TEMPLATE_PATH,
]


@pytest.mark.parametrize("path", REQUIRED_TEMPLATES, ids=lambda p: p.name)
def test_required_template_paths_resolve_to_existing_files(path: Path):
    assert path.exists(), f"Vorlage fehlt: {path}"


@pytest.mark.parametrize("path", REQUIRED_TEMPLATES, ids=lambda p: p.name)
def test_templates_can_be_opened_with_openpyxl(path: Path):
    workbook = openpyxl.load_workbook(path, read_only=True)
    try:
        assert workbook.sheetnames
    finally:
        workbook.close()


@pytest.mark.parametrize("path", REQUIRED_TEMPLATES, ids=lambda p: p.name)
def test_reading_a_template_does_not_modify_it(path: Path):
    before = path.read_bytes()
    workbook = openpyxl.load_workbook(path, read_only=True)
    _ = workbook.sheetnames
    workbook.close()
    after = path.read_bytes()

    assert before == after


def test_require_template_raises_a_clear_error_for_a_missing_file(tmp_path):
    missing = tmp_path / "does-not-exist.xlsx"

    with pytest.raises(FileNotFoundError, match="nicht gefunden"):
        paths_module.require_template(missing)


def test_require_template_returns_the_path_when_it_exists():
    result = paths_module.require_template(paths_module.WEEK_A_TEMPLATE_PATH)
    assert result == paths_module.WEEK_A_TEMPLATE_PATH
