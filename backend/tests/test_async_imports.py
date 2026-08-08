"""Tests für AP7 (blockierende Upload-/Import-Endpunkte aus dem Event-Loop lösen).

Deckt ab:
  - Schritt 8: während ein künstlich verlangsamter PDF-/XLSX-Import läuft,
    antwortet ein paralleler GET /api/health zeitnah (Event-Loop-Responsivität).
  - Schritt 9: Ergebnisgleichheit für PDF-, Probenplan-, XLSX- und
    Artist-Plan-Import (Response, Statuscodes, Fehlerpfade) - externe Dienste
    (Gemini) sind dabei immer gemockt, es werden keine echten/kostenpflichtigen
    Requests ausgeführt.
  - Schritt 10: Ressourcenfreigabe (Upload-Datei/Workbook) auch im Fehlerfall.

Immer gegen eine temporäre, per monkeypatch umgeleitete Testdatenbank - nie
gegen die echte lokale Mitarbeiterdatenbank.
"""
from __future__ import annotations

import io
import threading
import time

import openpyxl
import pytest
from fastapi.testclient import TestClient

from backend import api, db
from backend.routers import imports as imports_router


def _init_db(tmp_path, monkeypatch, filename: str):
    """AP4-Konvention: get_conn() legt kein Schema mehr an, deshalb hier einmalig
    explizit initialize_database() aufrufen."""


def _minimal_xlsx_bytes() -> bytes:
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


# ---------- Schritt 8: Event-Loop-Responsivität ----------


def test_health_responds_quickly_during_slow_pdf_import(tmp_path, monkeypatch):
    """Simuliert einen langsamen Gemini-Aufruf (0.4s synchrones Sleep) und prüft,
    dass ein paralleler /api/health-Request währenddessen zeitnah antwortet -
    der Beweis, dass upload_pdf den Event-Loop nicht mehr blockiert."""
    _init_db(tmp_path, monkeypatch, "responsiveness.db")

    def slow_extract(content, api_key=None):
        time.sleep(0.4)
        return {
            "kw": 1, "start_date": "2026-08-03", "end_date": "2026-08-09",
            "assignments": [], "absences": [],
        }

    monkeypatch.setattr(imports_router, "extract_dienstplan", slow_extract)

    results: dict[str, float] = {}
    import_started = threading.Event()

    with TestClient(api.app) as client:
        def run_import():
            import_started.set()
            start = time.perf_counter()
            resp = client.post(
                "/api/upload/pdf",
                params={"api_key": "test-key"},
                files={"file": ("plan.pdf", b"%PDF-1.4 fake", "application/pdf")},
            )
            results["import_status"] = resp.status_code
            results["import_elapsed"] = time.perf_counter() - start

        thread = threading.Thread(target=run_import)
        thread.start()
        import_started.wait(timeout=2)
        time.sleep(0.05)  # dem Import-Request Zeit geben, wirklich zu starten

        health_start = time.perf_counter()
        health_resp = client.get("/api/health")
        health_elapsed = time.perf_counter() - health_start

        thread.join(timeout=5)

    assert results["import_status"] == 200
    assert results["import_elapsed"] >= 0.4, "der künstliche Import muss tatsächlich ~0.4s dauern"
    assert health_resp.status_code == 200
    assert health_elapsed < 0.25, (
        f"/api/health durfte durch den parallelen Import nicht blockiert werden "
        f"(Laufzeit: {health_elapsed*1000:.0f} ms, Import: {results['import_elapsed']*1000:.0f} ms)"
    )


def test_health_responds_quickly_during_slow_xlsx_upload(tmp_path, monkeypatch):
    """Wie oben, aber für den XLSX-Upload-Pfad (openpyxl statt Gemini)."""
    _init_db(tmp_path, monkeypatch, "responsiveness-xlsx.db")

    from backend import xlsx_template

    real_extract = xlsx_template.extract_from_xlsx

    def slow_extract(path, sheet_name=None):
        time.sleep(0.4)
        return real_extract(path, sheet_name)

    monkeypatch.setattr(xlsx_template, "extract_from_xlsx", slow_extract)

    xlsx_bytes = _minimal_xlsx_bytes()
    results: dict[str, float] = {}
    import_started = threading.Event()

    with TestClient(api.app) as client:
        def run_import():
            import_started.set()
            start = time.perf_counter()
            resp = client.post(
                "/api/upload/xlsx",
                files={"file": ("plan.xlsx", xlsx_bytes, "application/vnd.openxmlformats")},
            )
            results["import_status"] = resp.status_code
            results["import_elapsed"] = time.perf_counter() - start

        thread = threading.Thread(target=run_import)
        thread.start()
        import_started.wait(timeout=2)
        time.sleep(0.05)

        health_start = time.perf_counter()
        health_resp = client.get("/api/health")
        health_elapsed = time.perf_counter() - health_start

        thread.join(timeout=5)

    # Die minimale XLSX hat keine erkennbare Datumszeile - erwartet ist ein
    # 500 ("Einlesen fehlgeschlagen"), das ist für diesen Test irrelevant: es
    # geht ausschließlich um die Event-Loop-Responsivität währenddessen.
    assert results["import_status"] == 500
    assert results["import_elapsed"] >= 0.4
    assert health_resp.status_code == 200
    assert health_elapsed < 0.25, (
        f"/api/health durfte durch den parallelen XLSX-Upload nicht blockiert werden "
        f"(Laufzeit: {health_elapsed*1000:.0f} ms, Import: {results['import_elapsed']*1000:.0f} ms)"
    )


# ---------- Schritt 9: Ergebnisgleichheit ----------


def test_pdf_import_success_returns_mocked_extraction(tmp_path, monkeypatch):
    _init_db(tmp_path, monkeypatch, "pdf-success.db")
    expected = {
        "kw": 32, "start_date": "2026-08-03", "end_date": "2026-08-09",
        "assignments": [
            {"date": "2026-08-03", "category": "Kochdienste", "subcategory": None,
             "person": "Tobi", "raw_text": "Tobi"}
        ],
        "absences": [],
    }
    monkeypatch.setattr(imports_router, "extract_dienstplan", lambda content, api_key=None: expected)

    with TestClient(api.app) as client:
        resp = client.post(
            "/api/upload/pdf", params={"api_key": "test-key"},
            files={"file": ("plan.pdf", b"%PDF-1.4 fake", "application/pdf")},
        )
    assert resp.status_code == 200
    assert resp.json() == expected


def test_pdf_import_missing_api_key_returns_400(tmp_path, monkeypatch):
    _init_db(tmp_path, monkeypatch, "pdf-missing-key.db")
    monkeypatch.delenv("GEMINI_API_KEY", raising=False)

    with TestClient(api.app) as client:
        resp = client.post(
            "/api/upload/pdf",
            files={"file": ("plan.pdf", b"%PDF-1.4 fake", "application/pdf")},
        )
    assert resp.status_code == 400
    assert resp.json()["detail"] == "Kein Gemini API Key vorhanden."


def test_pdf_import_extraction_error_returns_500(tmp_path, monkeypatch):
    _init_db(tmp_path, monkeypatch, "pdf-error.db")

    def raising_extract(content, api_key=None):
        raise RuntimeError("Gemini nicht erreichbar")

    monkeypatch.setattr(imports_router, "extract_dienstplan", raising_extract)

    with TestClient(api.app) as client:
        resp = client.post(
            "/api/upload/pdf", params={"api_key": "test-key"},
            files={"file": ("plan.pdf", b"%PDF-1.4 fake", "application/pdf")},
        )
    assert resp.status_code == 500
    assert "Extraktion fehlgeschlagen" in resp.json()["detail"]


def test_xlsx_sheets_lists_real_sheet_names(tmp_path, monkeypatch):
    _init_db(tmp_path, monkeypatch, "xlsx-sheets.db")
    xlsx_bytes = _minimal_xlsx_bytes()

    with TestClient(api.app) as client:
        resp = client.post(
            "/api/upload/xlsx/sheets",
            files={"file": ("plan.xlsx", xlsx_bytes, "application/vnd.openxmlformats")},
        )
    assert resp.status_code == 200
    assert resp.json() == {"sheets": ["Sheet1"]}


def test_xlsx_sheets_invalid_file_returns_500(tmp_path, monkeypatch):
    _init_db(tmp_path, monkeypatch, "xlsx-sheets-invalid.db")

    with TestClient(api.app) as client:
        resp = client.post(
            "/api/upload/xlsx/sheets",
            files={"file": ("plan.xlsx", b"not an xlsx file", "application/vnd.openxmlformats")},
        )
    assert resp.status_code == 500
    assert "Konnte Excel-Datei nicht lesen" in resp.json()["detail"]


def test_xlsx_import_invalid_file_returns_500(tmp_path, monkeypatch):
    _init_db(tmp_path, monkeypatch, "xlsx-import-invalid.db")

    with TestClient(api.app) as client:
        resp = client.post(
            "/api/upload/xlsx",
            files={"file": ("plan.xlsx", b"not an xlsx file", "application/vnd.openxmlformats")},
        )
    assert resp.status_code == 500
    assert "Einlesen fehlgeschlagen" in resp.json()["detail"]


def test_artist_plan_sheets_lists_real_sheet_names(tmp_path, monkeypatch):
    _init_db(tmp_path, monkeypatch, "artist-sheets.db")
    xlsx_bytes = _minimal_xlsx_bytes()

    with TestClient(api.app) as client:
        resp = client.post(
            "/api/artist-plans/upload/sheets",
            files={"file": ("plan.xlsx", xlsx_bytes, "application/vnd.openxmlformats")},
        )
    assert resp.status_code == 200
    assert resp.json() == {"sheets": ["Sheet1"]}


def test_artist_plan_import_success_with_minimal_workbook(tmp_path, monkeypatch):
    _init_db(tmp_path, monkeypatch, "artist-import-success.db")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KW32"
    ws.cell(2, 2).value = "03.08.2026"
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()

    with TestClient(api.app) as client:
        resp = client.post(
            "/api/artist-plans/import",
            files={"file": ("plan.xlsx", buf.getvalue(), "application/vnd.openxmlformats")},
        )
    assert resp.status_code == 200
    body = resp.json()
    assert body["sheet_name"] == "KW32"
    assert body["source_filename"] == "plan.xlsx"
    assert len(body["day_labels"]) == 7


def test_artist_plan_import_invalid_workbook_returns_400(tmp_path, monkeypatch):
    _init_db(tmp_path, monkeypatch, "artist-import-invalid.db")
    xlsx_bytes = _minimal_xlsx_bytes()  # kein Startdatum in B2

    with TestClient(api.app) as client:
        resp = client.post(
            "/api/artist-plans/import",
            files={"file": ("plan.xlsx", xlsx_bytes, "application/vnd.openxmlformats")},
        )
    assert resp.status_code == 400
    assert "Wochen-Startdatum" in resp.json()["detail"]


def test_rehearsal_plan_import_rejects_wrong_extension(tmp_path, monkeypatch):
    _init_db(tmp_path, monkeypatch, "rehearsal-wrong-ext.db")

    with TestClient(api.app) as client:
        resp = client.post(
            "/api/rehearsal-plans/import",
            files={"file": ("plan.docx", b"irrelevant", "application/msword")},
        )
    assert resp.status_code == 400
    assert resp.json()["detail"] == "Bitte eine Probenplan-PDF oder -Excel auswählen."


def test_rehearsal_plan_import_gemini_success_matches_mocked_result(tmp_path, monkeypatch):
    """Mockt rehearsal_plan.extract_pdf_with_gemini direkt - keine echten Gemini-Aufrufe."""
    _init_db(tmp_path, monkeypatch, "rehearsal-gemini.db")
    monkeypatch.setenv("GEMINI_API_KEY", "test-key")

    from backend import rehearsal_plan

    expected = {
        "start_date": "2026-08-03", "end_date": "2026-08-09",
        "source_filename": "probe.pdf", "rehearsals": [], "warnings": [],
        "extraction_method": "gemini",
    }
    monkeypatch.setattr(
        rehearsal_plan, "extract_pdf_with_gemini",
        lambda pdf_bytes, active_people, source_filename=None, api_key=None: expected,
    )

    with TestClient(api.app) as client:
        resp = client.post(
            "/api/rehearsal-plans/import",
            files={"file": ("probe.pdf", b"%PDF-1.4 fake", "application/pdf")},
        )
    assert resp.status_code == 200
    assert resp.json() == expected


def test_rehearsal_plan_import_falls_back_to_local_extraction_on_gemini_error(tmp_path, monkeypatch):
    """Schlägt Gemini fehl, greift der bestehende lokale Fallback - unverändertes
    Verhalten, nur jetzt im Threadpool statt im Event-Loop."""
    _init_db(tmp_path, monkeypatch, "rehearsal-gemini-fallback.db")
    monkeypatch.setenv("GEMINI_API_KEY", "test-key")

    from backend import rehearsal_plan

    def raising_gemini(pdf_bytes, active_people, source_filename=None, api_key=None):
        raise RuntimeError("Gemini nicht erreichbar")

    fallback_result = {
        "start_date": "2026-08-03", "end_date": "2026-08-09",
        "source_filename": "probe.pdf", "rehearsals": [], "warnings": [],
        "extraction_method": "local",
    }
    monkeypatch.setattr(rehearsal_plan, "extract_pdf_with_gemini", raising_gemini)
    monkeypatch.setattr(
        rehearsal_plan, "extract_pdf",
        lambda pdf_bytes, active_people, source_filename=None: fallback_result,
    )

    with TestClient(api.app) as client:
        resp = client.post(
            "/api/rehearsal-plans/import",
            files={"file": ("probe.pdf", b"%PDF-1.4 fake", "application/pdf")},
        )
    assert resp.status_code == 200
    body = resp.json()
    assert body["extraction_method"] == "local"
    assert body["warnings"][0].startswith("Gemini war gerade nicht verfügbar")


# ---------- Schritt 10: Ressourcenfreigabe (auch im Fehlerfall) ----------


def test_extract_from_xlsx_closes_workbook_even_on_error(monkeypatch):
    from backend import xlsx_template

    real_load_workbook = xlsx_template.openpyxl.load_workbook
    closed = {"n": 0}

    def spy_load_workbook(*args, **kwargs):
        wb = real_load_workbook(*args, **kwargs)
        real_close = wb.close

        def spy_close():
            closed["n"] += 1
            real_close()

        wb.close = spy_close
        return wb

    monkeypatch.setattr(xlsx_template.openpyxl, "load_workbook", spy_load_workbook)

    xlsx_bytes = _minimal_xlsx_bytes()  # keine Datumszeile -> ValueError im try-Block
    with pytest.raises(ValueError, match="Datumszeile"):
        xlsx_template.extract_from_xlsx(io.BytesIO(xlsx_bytes))

    assert closed["n"] == 1, "Workbook muss auch bei einer Exception geschlossen werden"


def test_artist_plan_extract_closes_workbook_even_on_error(monkeypatch):
    from backend import artist_plan

    real_load_workbook = artist_plan.load_workbook
    closed = {"n": 0}

    def spy_load_workbook(*args, **kwargs):
        wb = real_load_workbook(*args, **kwargs)
        real_close = wb.close

        def spy_close():
            closed["n"] += 1
            real_close()

        wb.close = spy_close
        return wb

    monkeypatch.setattr(artist_plan, "load_workbook", spy_load_workbook)

    xlsx_bytes = _minimal_xlsx_bytes()  # kein Startdatum in B2 -> ValueError
    with pytest.raises(ValueError, match="Wochen-Startdatum"):
        artist_plan.extract_artist_plan(xlsx_bytes)

    assert closed["n"] == 1, "Workbook muss auch bei einer Exception geschlossen werden"


def test_artist_plan_list_sheets_closes_workbook(monkeypatch):
    from backend import artist_plan

    real_load_workbook = artist_plan.load_workbook
    closed = {"n": 0}

    def spy_load_workbook(*args, **kwargs):
        wb = real_load_workbook(*args, **kwargs)
        real_close = wb.close

        def spy_close():
            closed["n"] += 1
            real_close()

        wb.close = spy_close
        return wb

    monkeypatch.setattr(artist_plan, "load_workbook", spy_load_workbook)

    sheets = artist_plan.list_sheets(io.BytesIO(_minimal_xlsx_bytes()))
    assert sheets == ["Sheet1"]
    assert closed["n"] == 1


def test_rehearsal_plan_extract_pdf_closes_document_even_on_error(monkeypatch):
    from backend import rehearsal_plan

    real_fitz_open = rehearsal_plan.fitz.open
    closed = {"n": 0}

    def spy_fitz_open(*args, **kwargs):
        document = real_fitz_open(*args, **kwargs)
        real_close = document.close

        def spy_close():
            closed["n"] += 1
            real_close()

        document.close = spy_close
        return document

    monkeypatch.setattr(rehearsal_plan.fitz, "open", spy_fitz_open)
    monkeypatch.setattr(
        rehearsal_plan, "_pdf_year",
        lambda document: (_ for _ in ()).throw(RuntimeError("kaputte PDF-Metadaten")),
    )

    # Ein minimales, aber strukturell gültiges PDF - fitz.open() muss erfolgreich
    # öffnen, damit der Fehler gezielt erst in _pdf_year() (innerhalb try/finally)
    # ausgelöst wird.
    minimal_pdf = (
        b"%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n"
        b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n"
        b"3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 200 200]>>endobj\n"
        b"trailer<</Root 1 0 R>>"
    )

    with pytest.raises(RuntimeError, match="kaputte PDF-Metadaten"):
        rehearsal_plan.extract_pdf(minimal_pdf, [])

    assert closed["n"] == 1, "PyMuPDF-Dokument muss auch bei einer Exception geschlossen werden"
