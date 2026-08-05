"""Import, Bearbeitung und Dienstplan-Übernahme des Künstlerplans."""
from __future__ import annotations

import io
import re
import unicodedata
from datetime import date, datetime, timedelta
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from . import db
from . import util

FIELD_DEFINITIONS = [
    {"key": "lunch_dj", "label": "Mittagsgrill · DJ", "group": "Tages-Entertainment"},
    {"key": "chillout_time", "label": "Chillout · Zeit", "group": "Tages-Entertainment"},
    {"key": "chillout_location", "label": "Chillout · Ort", "group": "Tages-Entertainment"},
    {"key": "chillout_artist", "label": "Chillout · Künstler/DJ", "group": "Tages-Entertainment"},
    {"key": "gastro_time", "label": "Gastrotainment · Zeit", "group": "Tages-Entertainment"},
    {"key": "gastro_location", "label": "Gastrotainment · Ort", "group": "Tages-Entertainment"},
    {"key": "gastro_artist", "label": "Gastrotainment · Künstler", "group": "Tages-Entertainment"},
    {"key": "evening_program", "label": "Abendprogramm · Show/Party", "group": "Abend-Entertainment"},
    {"key": "evening_dj", "label": "Schachbrett · DJ", "group": "Abend-Entertainment"},
    {"key": "special", "label": "Special", "group": "Zusatzinformationen"},
    {"key": "nite_club", "label": "NITE CLUB · DJ", "group": "Abend-Entertainment"},
]

FIELD_KEYS = {field["key"] for field in FIELD_DEFINITIONS}
LOCATION_WORDS = {
    "morroskai", "tennisbar", "beach", "wiese", "waspo", "eventfläche",
    "eventflache", "cofete", "vista mar", "poolbar", "schachbrett",
}
ARTIST_WORDS = {
    "dj", "djane", "live", "gesang", "band", "collection", "collective",
    "duo", "sax",
}


def _norm(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text).strip().casefold()


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


def _direct_cell_text(ws, row: int | None, column: int, *, allow_vertical_merge: bool = True) -> str:
    if row is None:
        return ""
    coordinate = ws.cell(row, column).coordinate
    for merged in ws.merged_cells.ranges:
        if coordinate in merged:
            if (
                not allow_vertical_merge
                and merged.max_row > merged.min_row
            ):
                return ""
            break
    return _cell_text(ws.cell(row, column).value)


def _find_label_row(ws, *needles: str) -> int | None:
    normalized_needles = [_norm(needle) for needle in needles]
    for row in range(1, ws.max_row + 1):
        value = _norm(ws.cell(row, 1).value)
        if any(needle in value for needle in normalized_needles):
            return row
    return None


def _start_date(ws) -> date:
    value = ws.cell(2, 2).value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for pattern in ("%d.%m.%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), pattern).date()
            except ValueError:
                pass
    raise ValueError(
        f"Im Reiter „{ws.title}“ wurde in B2 kein gültiges Wochen-Startdatum gefunden."
    )


def list_sheets(source: str | BinaryIO) -> list[str]:
    workbook = load_workbook(source, read_only=True, data_only=False)
    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def _block_values(ws, column: int, start_row: int | None, end_row: int) -> list[str]:
    if start_row is None:
        return []
    return [
        text
        for row in range(start_row, min(end_row, ws.max_row) + 1)
        if (text := _cell_text(ws.cell(row, column).value))
    ]


def _row_before_next(start_row: int | None, *anchors: int | None, fallback: int) -> int:
    if start_row is None:
        return fallback
    following = [row for row in anchors if row is not None and row > start_row]
    return min(following) - 1 if following else fallback


def _looks_like_time(text: str) -> bool:
    normalized = _norm(text)
    return bool(
        re.search(r"\d{1,2}\s*[:.]\s*\d{2}", normalized)
        or re.search(r"\d{1,2}\s*-\s*\d{1,2}\s*uhr", normalized)
    )


def _looks_like_artist(text: str) -> bool:
    normalized = _norm(text)
    return any(word in normalized for word in ARTIST_WORDS)


def _looks_like_location(text: str) -> bool:
    normalized = _norm(text)
    return (
        any(word in normalized for word in LOCATION_WORDS)
        or normalized.startswith(("ef ", "vm ", "ms "))
        or "apero" in normalized
    )


def _classify_details(values: list[str]) -> tuple[str, str, str]:
    times: list[str] = []
    locations: list[str] = []
    artists: list[str] = []
    leftovers: list[str] = []
    for value in values:
        if _looks_like_artist(value):
            artists.append(value)
        elif _looks_like_time(value):
            times.append(value)
        elif _looks_like_location(value):
            locations.append(value)
        elif value not in {"-", "–"}:
            leftovers.append(value)
    # Unbekannte Namen ohne DJ-/LIVE-Zusatz sind wahrscheinlicher Künstler als Orte.
    artists.extend(leftovers)
    return "\n".join(times), "\n".join(locations), "\n".join(artists)


def empty_rows(start_date: date) -> dict:
    week_dates = [(start_date + timedelta(days=offset)).isoformat() for offset in range(7)]
    day_labels = [util.fmt_date_short(day) for day in week_dates]
    rows = []
    for field in FIELD_DEFINITIONS:
        row = {
            "field_key": field["key"],
            "label": field["label"],
            "group": field["group"],
        }
        row.update({label: "" for label in day_labels})
        rows.append(row)
    return {
        "start_date": start_date.isoformat(),
        "end_date": (start_date + timedelta(days=6)).isoformat(),
        "week_dates_iso": week_dates,
        "day_labels": day_labels,
        "rows": rows,
    }


def extract_artist_plan(
    source: bytes | BinaryIO,
    sheet_name: str | None = None,
    source_filename: str | None = None,
) -> dict:
    file_source = io.BytesIO(source) if isinstance(source, bytes) else source
    workbook = load_workbook(file_source, data_only=False)
    try:
        ws = workbook[sheet_name] if sheet_name else workbook.active
        start = _start_date(ws)
        result = empty_rows(start)
        result.update({
            "sheet_name": ws.title,
            "source_filename": source_filename,
        })

        lunch_row = _find_label_row(ws, "Mittagsgrill")
        chillout_row = _find_label_row(ws, "Chillout")
        gastro_row = _find_label_row(ws, "Gastrotainment")
        evening_row = _find_label_row(ws, "Abend Programm", "Programm Abend")
        schachbrett_row = _find_label_row(ws, "Schachbrett")
        special_row = _find_label_row(ws, "Special")
        nite_row = _find_label_row(ws, "Nite Club")

        row_lookup = {row["field_key"]: row for row in result["rows"]}
        for day_index, column in enumerate(range(2, 9)):
            day_label = result["day_labels"][day_index]
            row_lookup["lunch_dj"][day_label] = _direct_cell_text(
                ws, lunch_row, column, allow_vertical_merge=False
            )

            chill_values = _block_values(
                ws,
                column,
                chillout_row,
                _row_before_next(
                    chillout_row,
                    gastro_row,
                    evening_row,
                    schachbrett_row,
                    fallback=ws.max_row,
                ),
            )
            chill_time, chill_location, chill_artist = _classify_details(chill_values)
            row_lookup["chillout_time"][day_label] = chill_time
            row_lookup["chillout_location"][day_label] = chill_location
            row_lookup["chillout_artist"][day_label] = chill_artist

            gastro_values = _block_values(
                ws,
                column,
                gastro_row,
                _row_before_next(
                    gastro_row,
                    evening_row,
                    schachbrett_row,
                    special_row,
                    fallback=ws.max_row,
                ),
            )
            gastro_time, gastro_location, gastro_artist = _classify_details(gastro_values)
            row_lookup["gastro_time"][day_label] = gastro_time
            row_lookup["gastro_location"][day_label] = gastro_location
            row_lookup["gastro_artist"][day_label] = gastro_artist

            row_lookup["evening_program"][day_label] = _direct_cell_text(
                ws, evening_row, column
            )
            row_lookup["evening_dj"][day_label] = _direct_cell_text(
                ws, schachbrett_row, column
            )
            row_lookup["special"][day_label] = _direct_cell_text(
                ws, special_row, column
            )
            row_lookup["nite_club"][day_label] = _direct_cell_text(
                ws, nite_row, column
            )
        return result
    finally:
        workbook.close()


def rows_to_entries(rows: list[dict], day_labels: list[str], week_dates_iso: list[str]) -> list[dict]:
    entries = []
    for row in rows:
        field_key = row.get("field_key")
        if field_key not in FIELD_KEYS:
            continue
        for label, date_iso in zip(day_labels, week_dates_iso):
            entries.append({
                "date": date_iso,
                "field_key": field_key,
                "content": str(row.get(label) or "").strip(),
            })
    return entries


def stored_plan_payload(conn, plan_row) -> dict:
    result = empty_rows(datetime.strptime(plan_row["start_date"], "%Y-%m-%d").date())
    day_by_date = dict(zip(result["week_dates_iso"], result["day_labels"]))
    rows_by_key = {row["field_key"]: row for row in result["rows"]}
    for entry in db.get_artist_plan_entries(conn, plan_row["id"]):
        day_label = day_by_date.get(entry["date"])
        target = rows_by_key.get(entry["field_key"])
        if target is not None and day_label:
            target[day_label] = entry["content"] or ""
    result.update({
        "id": plan_row["id"],
        "source_filename": plan_row["source_filename"],
        "sheet_name": plan_row["sheet_name"],
    })
    return result


def _entry_map(conn, artist_plan_id: int) -> dict[tuple[str, str], str]:
    return {
        (row["date"], row["field_key"]): (row["content"] or "").strip()
        for row in db.get_artist_plan_entries(conn, artist_plan_id)
    }


def _joined(values: list[str]) -> str:
    return "\n".join(value for value in values if value).strip()


def apply_to_draft(conn, draft: list[dict], plan_row) -> list[dict]:
    """Überschreibt nur künstlerbezogene Inhalte; interne MA-Zuweisungen bleiben erhalten."""
    entries = _entry_map(conn, plan_row["id"])
    dates = sorted({date_iso for date_iso, _ in entries})
    content_by_date: dict[str, dict[str, str]] = {}
    for date_iso in dates:
        get = lambda key: entries.get((date_iso, key), "")
        content_by_date[date_iso] = {
            "Mittagsgrill": get("lunch_dj"),
            "ChillOut Künstler": _joined([
                get("chillout_location"),
                get("chillout_time"),
                get("chillout_artist"),
            ]),
            "Show + Party": _joined([get("evening_program"), get("evening_dj")]),
            "Specials": get("special"),
            "NITE CLUB": get("nite_club"),
            "Aperitif": _joined([
                get("gastro_location"),
                get("gastro_time"),
                get("gastro_artist"),
            ]),
        }

    content_categories = {
        "Mittagsgrill", "ChillOut Künstler", "Show + Party", "Specials", "NITE CLUB",
    }
    next_draft = []
    aperitif_dates_done: set[str] = set()
    for row in draft:
        date_iso = row.get("date")
        category = row.get("category")
        replacement = content_by_date.get(date_iso, {}).get(category, "")
        if category in content_categories and replacement:
            continue
        if category == "Aperitif" and replacement:
            # Künstlerinfos bilden den Präfix; der intern zugeteilte MA bleibt hinter dem Trenner.
            if date_iso not in aperitif_dates_done:
                row = {**row, "subcategory": replacement}
                aperitif_dates_done.add(date_iso)
            else:
                row = {**row, "subcategory": None}
        next_draft.append(row)

    for date_iso, contents in content_by_date.items():
        for category in content_categories:
            text = contents.get(category, "")
            if text:
                next_draft.append({
                    "date": date_iso,
                    "category": category,
                    "subcategory": None,
                    "person": None,
                    "raw_text": text,
                })
        aperitif_text = contents.get("Aperitif", "")
        if aperitif_text and date_iso not in aperitif_dates_done:
            next_draft.append({
                "date": date_iso,
                "category": "Aperitif",
                "subcategory": aperitif_text,
                "person": None,
                "raw_text": aperitif_text,
            })
    return next_draft


def _detail_row_map(ws, anchor: int | None, end_row: int) -> dict[str, int]:
    if anchor is None:
        return {}
    result: dict[str, int] = {}
    for row in range(anchor, min(end_row, ws.max_row) + 1):
        values = [
            _cell_text(ws.cell(row, column).value)
            for column in range(2, 9)
            if _cell_text(ws.cell(row, column).value)
        ]
        text = "\n".join(values)
        if not text:
            continue
        if _looks_like_artist(text):
            result.setdefault("artist", row)
        elif _looks_like_time(text):
            result.setdefault("time", row)
        elif _looks_like_location(text):
            result.setdefault("location", row)
    if {"time", "location", "artist"} - result.keys():
        first_text = "\n".join(
            _cell_text(ws.cell(anchor, column).value)
            for column in range(2, 9)
        )
        if _looks_like_time(first_text):
            defaults = {"time": anchor, "location": anchor + 1, "artist": anchor + 2}
        else:
            defaults = {"location": anchor, "artist": anchor + 1, "time": anchor + 2}
        for key, row in defaults.items():
            result.setdefault(key, min(row, end_row))
    return result


def _write_cell(ws, row: int | None, column: int, value: str | date | None) -> None:
    if row is None:
        return
    cell = ws.cell(row, column)
    if isinstance(cell, MergedCell):
        return
    cell.value = value if value not in ("", None) else None


def export_artist_plan(
    conn,
    plan_row,
    template_path: str,
    output_path: str,
) -> str:
    """Schreibt eine gespeicherte Woche zurück in das echte Künstlerplan-Excel-Layout."""
    workbook = load_workbook(template_path)
    source_sheet = plan_row["sheet_name"]
    if source_sheet not in workbook.sheetnames:
        # Für neue Wochen das jüngste reguläre 7-Tage-Blatt verwenden. Großflächig
        # verbundene Sonderwochen eignen sich nicht als leere Standardvorlage.
        source_sheet = next(
            (
                ws.title
                for ws in reversed(workbook.worksheets)
                if not any(
                    merged.max_col - merged.min_col >= 1
                    and merged.max_row - merged.min_row >= 5
                    for merged in ws.merged_cells.ranges
                )
            ),
            workbook.sheetnames[-1],
        )
    ws = workbook[source_sheet]
    for other in list(workbook.worksheets):
        if other is not ws:
            workbook.remove(other)

    start = datetime.strptime(plan_row["start_date"], "%Y-%m-%d").date()
    entries = _entry_map(conn, plan_row["id"])
    for day_index, column in enumerate(range(2, 9)):
        _write_cell(ws, 2, column, start + timedelta(days=day_index))

    lunch_row = _find_label_row(ws, "Mittagsgrill")
    chillout_row = _find_label_row(ws, "Chillout")
    gastro_row = _find_label_row(ws, "Gastrotainment")
    evening_row = _find_label_row(ws, "Abend Programm", "Programm Abend")
    schachbrett_row = _find_label_row(ws, "Schachbrett")
    special_row = _find_label_row(ws, "Special")
    nite_row = _find_label_row(ws, "Nite Club")
    chill_end = _row_before_next(
        chillout_row, gastro_row, evening_row, schachbrett_row, fallback=ws.max_row
    )
    gastro_end = _row_before_next(
        gastro_row, evening_row, schachbrett_row, special_row, fallback=ws.max_row
    )
    chill_rows = _detail_row_map(ws, chillout_row, chill_end)
    gastro_rows = _detail_row_map(ws, gastro_row, gastro_end)

    # Nur die programmrelevanten Zellen leeren; Formatierungen und verbundene Zellen bleiben.
    editable_rows = {
        row
        for row in [
            lunch_row,
            evening_row,
            schachbrett_row,
            special_row,
            nite_row,
            *chill_rows.values(),
            *gastro_rows.values(),
        ]
        if row is not None
    }
    for row in editable_rows:
        for column in range(2, 9):
            _write_cell(ws, row, column, None)

    for day_index, column in enumerate(range(2, 9)):
        date_iso = (start + timedelta(days=day_index)).isoformat()
        get = lambda key: entries.get((date_iso, key), "")
        _write_cell(ws, lunch_row, column, get("lunch_dj"))
        _write_cell(ws, chill_rows.get("time"), column, get("chillout_time"))
        _write_cell(ws, chill_rows.get("location"), column, get("chillout_location"))
        _write_cell(ws, chill_rows.get("artist"), column, get("chillout_artist"))
        _write_cell(ws, gastro_rows.get("time"), column, get("gastro_time"))
        _write_cell(ws, gastro_rows.get("location"), column, get("gastro_location"))
        _write_cell(ws, gastro_rows.get("artist"), column, get("gastro_artist"))
        _write_cell(ws, evening_row, column, get("evening_program"))
        _write_cell(ws, schachbrett_row, column, get("evening_dj"))
        _write_cell(ws, special_row, column, get("special"))
        _write_cell(ws, nite_row, column, get("nite_club"))

    ws.title = f"KW {start.isocalendar()[1]}_{start.strftime('%d.%m')}"[:31]
    workbook.save(output_path)
    workbook.close()
    return output_path
